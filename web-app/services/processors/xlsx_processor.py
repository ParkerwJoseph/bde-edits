import time
from typing import List

from config.settings import LLM_MAX_INPUT_TOKENS
from utils.logger import get_logger

logger = get_logger(__name__)

# Approximate chars per token for estimation
CHARS_PER_TOKEN = 4


class XLSXProcessor:
    """
    Processes XLSX (Excel) documents by extracting text content from sheets.
    Splits content into batches respecting token limits.
    """

    def __init__(self):
        logger.info("[XLSXProcessor] Initialized")

    def process(self, file_path: str) -> List[dict]:
        """
        Process XLSX file and extract text content.

        Args:
            file_path: Path to the XLSX file

        Returns:
            List of section dicts with 'page_number', 'content_type', 'text_content', etc.
        """
        logger.info(f"[XLSXProcessor] Processing XLSX...")
        start_time = time.time()

        try:
            import openpyxl
        except ImportError:
            logger.error("[XLSXProcessor] openpyxl not installed")
            return [{
                "page_number": 1,
                "content_type": "text",
                "text_content": "[Excel file - install openpyxl for extraction]",
                "char_count": 45,
                "estimated_tokens": 12
            }]

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sections = []

        # Calculate safe token limit
        safe_token_limit = LLM_MAX_INPUT_TOKENS - 2000

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_sections = self._process_sheet(sheet, sheet_name, safe_token_limit)
            sections.extend(sheet_sections)

        wb.close()

        # Renumber sections sequentially
        for idx, section in enumerate(sections, start=1):
            section["page_number"] = idx

        # If no content extracted
        if not sections:
            sections.append({
                "page_number": 1,
                "content_type": "text",
                "text_content": "[Empty spreadsheet]",
                "char_count": 19,
                "estimated_tokens": 5
            })

        elapsed = time.time() - start_time
        logger.info(f"[XLSXProcessor] Done in {elapsed:.2f}s - {len(sections)} sections")

        return sections

    def _process_sheet(self, sheet, sheet_name: str, token_limit: int) -> List[dict]:
        """
        Process a single sheet and split into token-safe sections.

        Args:
            sheet: openpyxl worksheet
            sheet_name: Name of the sheet
            token_limit: Maximum tokens per section

        Returns:
            List of section dicts
        """
        sections = []
        current_rows = []
        current_tokens = 0
        header = f"SHEET: {sheet_name}\n"
        header_tokens = len(header) // CHARS_PER_TOKEN

        def flush_section():
            nonlocal current_rows, current_tokens
            if current_rows:
                text = header + "\n".join(current_rows)
                sections.append({
                    "page_number": 0,  # Will be renumbered later
                    "content_type": "text",
                    "text_content": text,
                    "char_count": len(text),
                    "estimated_tokens": len(text) // CHARS_PER_TOKEN,
                    "sheet_name": sheet_name
                })
                logger.info(f"  Sheet '{sheet_name}' section: {len(text)} chars, ~{len(text) // CHARS_PER_TOKEN} tokens")
                current_rows = []
                current_tokens = header_tokens

        current_tokens = header_tokens

        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell) if cell is not None else "" for cell in row]

            # Skip empty rows
            if not any(v.strip() for v in row_values):
                continue

            row_text = " | ".join(row_values)
            row_tokens = len(row_text) // CHARS_PER_TOKEN + 1

            # Check if adding this row would exceed limit
            if current_tokens + row_tokens > token_limit:
                flush_section()

            current_rows.append(row_text)
            current_tokens += row_tokens

        # Flush remaining content
        flush_section()

        return sections
