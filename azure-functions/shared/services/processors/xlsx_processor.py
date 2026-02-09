import time
from typing import List

from shared.config.settings import LLM_MAX_INPUT_TOKENS
from shared.utils.logger import get_logger

logger = get_logger(__name__)

CHARS_PER_TOKEN = 4


class XLSXProcessor:
    """
    Processes XLSX (Excel) documents by extracting text content from sheets.
    """

    def __init__(self):
        logger.info("[XLSXProcessor] Initialized")

    def process(self, file_path: str) -> List[dict]:
        logger.info(f"[XLSXProcessor] Processing XLSX...")
        start_time = time.time()

        try:
            import openpyxl
        except ImportError:
            logger.error("[XLSXProcessor] openpyxl not installed")
            return [{
                "page_number": 1,
                "content_type": "text",
                "text_content": "[Excel file - install openpyxl for extraction]",
                "char_count": 45,
                "estimated_tokens": 12
            }]

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sections = []
        safe_token_limit = LLM_MAX_INPUT_TOKENS - 2000

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_sections = self._process_sheet(sheet, sheet_name, safe_token_limit)
            sections.extend(sheet_sections)

        wb.close()

        for idx, section in enumerate(sections, start=1):
            section["page_number"] = idx

        if not sections:
            sections.append({
                "page_number": 1,
                "content_type": "text",
                "text_content": "[Empty spreadsheet]",
                "char_count": 19,
                "estimated_tokens": 5
            })

        elapsed = time.time() - start_time
        logger.info(f"[XLSXProcessor] Done in {elapsed:.2f}s - {len(sections)} sections")

        return sections

    def _process_sheet(self, sheet, sheet_name: str, token_limit: int) -> List[dict]:
        sections = []
        current_rows = []
        current_tokens = 0
        header = f"SHEET: {sheet_name}\n"
        header_tokens = len(header) // CHARS_PER_TOKEN

        def flush_section():
            nonlocal current_rows, current_tokens
            if current_rows:
                text = header + "\n".join(current_rows)
                sections.append({
                    "page_number": 0,
                    "content_type": "text",
                    "text_content": text,
                    "char_count": len(text),
                    "estimated_tokens": len(text) // CHARS_PER_TOKEN,
                    "sheet_name": sheet_name
                })
                current_rows = []
                current_tokens = header_tokens

        current_tokens = header_tokens

        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell) if cell is not None else "" for cell in row]

            if not any(v.strip() for v in row_values):
                continue

            row_text = " | ".join(row_values)
            row_tokens = len(row_text) // CHARS_PER_TOKEN + 1

            if current_tokens + row_tokens > token_limit:
                flush_section()

            current_rows.append(row_text)
            current_tokens += row_tokens

        flush_section()

        return sections
